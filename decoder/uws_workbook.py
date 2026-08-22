"""THE FILTERABLE COMP SHEET — 110 West 88th Street.

    python uws_workbook.py

The operator's ask, exactly: "my filters are sf, units, built, and distance ...
just give me the spreadsheet with filter applicable". So every one of those four
is a NUMERIC COLUMN with an autofilter on it, not a cutoff applied here. A cutoff
I choose is a decision taken away from the person who knows the deal.

⚠ THE ROW IS A PARCEL, NOT A LISTING PAGE. StreetEasy splits one property into
    several building pages when it has several addresses or towers, and those
    pages share a tax lot. Rolling up to the BBL is the rule the operator set on
    Gotham Point (3 pages, 1 lot -> one comp) versus Malt Drive (2 pages, 2 lots
    -> two comps). `se_pages` records how many pages folded in, so a rolled-up
    row can always be told from a simple one.

⚠ RENT $/SF IS QUOTED ANNUAL, and the SF beneath it is a MEASURED MINORITY. Only
    a quarter of StreetEasy listings publish square footage, so every $/sf figure
    here carries `leases_with_sf` beside it. A median $/sf computed from two
    listings and one from two hundred cannot be read the same way, and without
    the denominator on the same row they look identical.

⚠ AND `sf_source` SAYS WHOSE SQUARE FOOTAGE IT IS. `bldg_gross_sf` is PLUTO's
    whole-building figure — the filter the operator wants for scale — while the
    per-listing sf is the apartment. They are different quantities in the same
    unit and must never share a column.
"""
import json, math, pathlib, statistics, sys, time
from collections import Counter, defaultdict

sys.path.insert(0, str(pathlib.Path(__file__).parent))
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HERE = pathlib.Path(__file__).parent
SCOPE = HERE / "uws_scope.json"
PLUTO = HERE / "uws_pluto.json"
OUT = HERE / "110 West 88th - Upper West Side Rental Comps.xlsx"
RECENT_FROM = "2023-01-01"          # the 3-year window, as a live comparison
PULLED_AT = time.strftime("%Y-%m-%d")


def num(v):
    try:
        f = float(v)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def med(xs):
    xs = [x for x in xs if x]
    return round(statistics.median(xs), 2) if xs else None


def harvest(want):
    """Every listing on these parcels, straight out of the raw pull."""
    rows = []
    for fp in sorted((HERE / "leases_raw").glob("leases_*.jsonl")):
        with open(fp, encoding="utf-8") as f:
            for line in f:
                i = line.find('"bbl":"')
                if i < 0 or line[i + 7:line.find('"', i + 7)] not in want:
                    continue
                try:
                    rows.append(json.loads(line))
                except Exception:
                    pass
    # ⚠ the same listing can appear in two pull files (a resumed run re-reads a
    # building). `source_id` is the identity — dedupe on it or every median is
    # computed over duplicated evidence.
    seen, out = set(), []
    for r in rows:
        k = (r.get("bbl"), str(r.get("source_id")))
        if k in seen:
            continue
        seen.add(k)
        out.append(r)
    return out


def build():
    scope = json.loads(SCOPE.read_text(encoding="utf-8"))
    pl = json.loads(PLUTO.read_text(encoding="utf-8"))
    subj = scope["subject"]
    bldgs = scope["buildings"]

    by_bbl = defaultdict(list)
    for b in bldgs:
        by_bbl[b["bbl"]].append(b)
    print(f"{len(bldgs):,} building pages -> {len(by_bbl):,} parcels")

    print("reading the harvest...")
    leases = harvest(set(by_bbl))
    print(f"  {len(leases):,} distinct listings")

    per = defaultdict(list)
    for r in leases:
        per[r["bbl"]].append(r)

    rows, lease_rows = [], []
    for bbl, group in by_bbl.items():
        group.sort(key=lambda g: g["miles"])
        p = pl.get(bbl) or {}
        ls = per.get(bbl, [])
        act = [r for r in ls if r.get("lane") == "active"]
        his = [r for r in ls if r.get("lane") != "active"]
        recent = [r for r in ls if (r.get("event_date") or "") >= RECENT_FROM]

        def stats(rs):
            amt = [num(r.get("amount")) for r in rs]
            sf = [(num(r.get("amount")), num(r.get("sf"))) for r in rs]
            psf = [a * 12 / s for a, s in sf if a and s and s > 50]
            return med(amt), med([s for _, s in sf if s]), med(psf), len(psf)

        a_rent, _, _, _ = stats(act)
        h_rent, h_sf, h_psf, h_n = stats(his)
        r_rent, r_sf, r_psf, r_n = stats(recent)
        allsf = sum(1 for r in ls if num(r.get("sf")))

        alter = max(num(p.get("yearalter1")) or 0, num(p.get("yearalter2")) or 0)
        built = num(p.get("yearbuilt")) or None
        se_built = next((g["se_built"] for g in group if g.get("se_built")), None)
        dates = sorted(r.get("event_date") or "" for r in ls if r.get("event_date"))

        rows.append({
            "distance_mi": group[0]["miles"],
            "building": " / ".join(dict.fromkeys(g["name"] for g in group if g.get("name"))),
            "address": p.get("address") or group[0].get("address"),
            "bbl": bbl,
            "bldg_gross_sf": num(p.get("bldgarea")),
            "resid_sf": num(p.get("resarea")),
            "comm_sf": num(p.get("comarea")),
            "lot_sf": num(p.get("lotarea")),
            "units_res_pluto": num(p.get("unitsres")),
            "units_total_pluto": num(p.get("unitstotal")),
            "units_streeteasy": next((g["se_units"] for g in group if g.get("se_units")), None),
            "stories": num(p.get("numfloors")),
            "year_built": built if built else None,
            "year_altered": alter or None,
            "year_built_or_alt": max(built or 0, alter or 0) or None,
            "se_year_built": se_built,
            "bldg_class": p.get("bldgclass"),
            "zoning": p.get("zonedist1"),
            "built_far": num(p.get("builtfar")),
            "resid_far": num(p.get("residfar")),
            "owner": p.get("ownername"),
            "zip": p.get("zipcode"),
            "listings_active": len(act),
            "listings_historical": len(his),
            "listings_total": len(ls),
            "leases_with_sf": allsf,
            "sf_coverage_pct": round(allsf / len(ls) * 100, 1) if ls else None,
            "med_rent_active": a_rent,
            "med_rent_historical": h_rent,
            f"med_rent_since_{RECENT_FROM[:4]}": r_rent,
            f"med_unit_sf_since_{RECENT_FROM[:4]}": r_sf,
            f"med_psf_yr_since_{RECENT_FROM[:4]}": r_psf,
            f"n_psf_since_{RECENT_FROM[:4]}": r_n,
            "med_psf_yr_all": h_psf,
            "n_psf_all": h_n,
            "first_seen": dates[0] if dates else None,
            "last_seen": dates[-1] if dates else None,
            "se_pages": len(group),
            "harvested": "yes" if ls else "NOT PULLED",
            "key_verdict": group[0].get("verdict"),
            "slug": group[0]["slug"],
            "lat": group[0]["lat"], "lon": group[0]["lon"],
            "pulled_at": PULLED_AT,
        })

        for r in ls:
            a, s = num(r.get("amount")), num(r.get("sf"))
            lease_rows.append({
                "distance_mi": group[0]["miles"],
                "bbl": bbl,
                "building": r.get("building_name"),
                "bldg_gross_sf": num(p.get("bldgarea")),
                "year_built": built if built else None,
                "year_altered": alter or None,
                "unit": r.get("unit"),
                "unit_type": r.get("unit_type"),
                "beds": r.get("beds"),
                "sf": s,
                "amount": a,
                "psf_yr": round(a * 12 / s, 2) if a and s and s > 50 else None,
                "status": r.get("lane"),
                "se_status": r.get("se_status"),
                "event_date": r.get("event_date"),
                "source_id": r.get("source_id"),
                "pulled_at": PULLED_AT,
            })

    rows.sort(key=lambda r: r["distance_mi"])
    lease_rows.sort(key=lambda r: (r["distance_mi"], r["bbl"],
                                   r["event_date"] or ""), reverse=False)
    return subj, rows, lease_rows


def write(subj, rows, lease_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_f = Font(bold=True, color="FFFFFF", size=9)
    head_b = PatternFill("solid", fgColor="1F4E3D")

    def sheet(ws, data, cols, widths=None, freeze="A2"):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = head_f, head_b
            cell.alignment = Alignment(wrap_text=True, vertical="bottom")
        for r in data:
            ws.append([r.get(c) for c in cols])
        ws.freeze_panes = freeze
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data)+1}"
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = \
                (widths or {}).get(c, min(max(10, len(c) + 2), 26))
        ws.row_dimensions[1].height = 30

    # ── the sheet the filters live on ───────────────────────────────────────
    ws = wb.active
    ws.title = "Buildings"
    cols = ["distance_mi", "building", "address", "bbl", "bldg_gross_sf",
            "units_res_pluto", "units_streeteasy", "year_built", "year_altered",
            "year_built_or_alt", "stories", "listings_active",
            "listings_historical", "listings_total", "leases_with_sf",
            "sf_coverage_pct", "med_rent_active",
            f"med_rent_since_{RECENT_FROM[:4]}",
            f"med_unit_sf_since_{RECENT_FROM[:4]}",
            f"med_psf_yr_since_{RECENT_FROM[:4]}",
            f"n_psf_since_{RECENT_FROM[:4]}",
            "med_rent_historical", "med_psf_yr_all", "n_psf_all",
            "resid_sf", "comm_sf", "lot_sf", "units_total_pluto", "bldg_class",
            "zoning", "built_far", "resid_far", "owner", "zip",
            "first_seen", "last_seen", "se_pages", "harvested", "key_verdict",
            "slug", "pulled_at"]
    sheet(ws, rows, cols, {"building": 34, "address": 26, "owner": 30, "slug": 30})
    for r in range(2, len(rows) + 2):
        ws.cell(row=r, column=1).number_format = "0.00"
        for c in (5, 6, 7, 25, 26, 27):
            ws.cell(row=r, column=c).number_format = "#,##0"
        for c in (17, 18, 22):
            ws.cell(row=r, column=c).number_format = "$#,##0"
        for c in (20, 23):
            ws.cell(row=r, column=c).number_format = "$#,##0.00"

    # ── every listing behind those numbers ──────────────────────────────────
    ws2 = wb.create_sheet("Leases")
    lcols = ["distance_mi", "bbl", "building", "bldg_gross_sf", "year_built",
             "year_altered", "unit", "unit_type", "beds", "sf", "amount",
             "psf_yr", "status", "se_status", "event_date", "source_id",
             "pulled_at"]
    sheet(ws2, lease_rows, lcols, {"building": 32})
    for r in range(2, len(lease_rows) + 2):
        ws2.cell(row=r, column=1).number_format = "0.00"
        ws2.cell(row=r, column=4).number_format = "#,##0"
        ws2.cell(row=r, column=10).number_format = "#,##0"
        ws2.cell(row=r, column=11).number_format = "$#,##0"
        ws2.cell(row=r, column=12).number_format = "$#,##0.00"

    # ── what the numbers are and are not ────────────────────────────────────
    ws3 = wb.create_sheet("Read me")
    pulled = [r for r in rows if r["harvested"] == "yes"]
    sf_rows = [r for r in rows if r["n_psf_all"]]
    note = [
        ("SUBJECT", f"{subj['label']} · BBL {subj['bbl']} · Block 1218 Lot 129"),
        ("", "PLUTO: 1,253 sf lot · class V1 (vacant) · R7-2 / C1-9 split · "
             "residential FAR 3.44, affordable-housing FAR 5.01"),
        ("", ""),
        ("SCOPE", f"THE UPPER WEST SIDE ONLY — Manhattan Community District 7, 59th to "
                  f"110th, Central Park West to the Hudson. {len(rows):,} parcels "
                  f"carrying {sum(r['se_pages'] for r in rows):,} StreetEasy rental "
                  f"building pages."),
        ("", "The boundary is the district line, not a circle. A radius drawn from 88th "
             "Street crosses the park into the Upper East Side and runs south into "
             "Hell's Kitchen, which is why the earlier version was wrong."),
        ("", "Distance from the subject is still a COLUMN — filter it to tighten to "
             "Manhattan Valley, the 80s, or Lincoln Square as the deal needs."),
        ("", ""),
        ("ROW = PARCEL", "StreetEasy splits one property into several building pages "
                         "when it has several addresses or towers. Those pages share a "
                         "tax lot and are folded into one row; se_pages says how many."),
        ("", ""),
        ("LEASE DATA", f"{len(pulled):,} of {len(rows):,} parcels "
                       f"({len(pulled)/len(rows)*100:.0f}%) have listings harvested; "
                       f"{len(rows)-len(pulled):,} are marked NOT PULLED and their "
                       f"rent columns are blank — blank means unknown, not zero."),
        ("", f"{len(lease_rows):,} listings on the Leases tab, active and historical."),
        ("", ""),
        ("⚠ SQUARE FOOTAGE", "bldg_gross_sf is PLUTO's whole-building area — the scale "
                             "filter. The sf column on the Leases tab is the APARTMENT. "
                             "Same unit, different quantity; never mix them."),
        ("", f"Only {sum(r['leases_with_sf'] for r in rows):,} of {len(lease_rows):,} "
             f"listings publish apartment square footage "
             f"({sum(r['leases_with_sf'] for r in rows)/max(len(lease_rows),1)*100:.0f}%). "
             f"That is why every $/sf column has its count (n_psf) beside it."),
        ("", f"{len(sf_rows):,} parcels have at least one SF-supported listing — those "
             f"are the ones that can carry a $/sf comp."),
        ("", ""),
        ("⚠ $/SF IS ANNUAL", "psf_yr = monthly rent x 12 / apartment sf, the way rental "
                             "comps are quoted. Divide by 12 for monthly."),
        ("", ""),
        ("⚠ ACTIVE vs HISTORICAL", "Active is what is on the market right now (an ASK). "
                                   "Historical is what came off the market — mostly "
                                   "leased, but se_status distinguishes RENTED from "
                                   "DELISTED and NO_LONGER_AVAILABLE."),
        ("", ""),
        ("YEAR", "year_built and year_altered are separate filters on purpose. A pre-war "
                 "building gut-renovated in 2019 competes with new construction; "
                 "year_built_or_alt takes the later of the two."),
        ("", ""),
        ("PULLED", f"StreetEasy listings harvested through {PULLED_AT}. "
                   f"PLUTO 26v1. Every row carries pulled_at."),
    ]
    ws3.append(["", ""])
    for a, b in note:
        ws3.append([a, b])
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 108
    for r in range(1, len(note) + 2):
        ws3.cell(row=r, column=1).font = Font(bold=True, size=9)
        ws3.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    subj, rows, lease_rows = build()
    print(f"\nwriting {len(rows):,} parcel rows and {len(lease_rows):,} listings...")
    p = write(subj, rows, lease_rows)
    print(f"wrote {p}  ({p.stat().st_size/1e6:.1f} MB)")
